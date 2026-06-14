#!/usr/bin/env python3

import os
import re
import sys
import smtplib
import subprocess
from pathlib import Path
from datetime import datetime
from email.message import EmailMessage

import requests
from flask import Flask, request, jsonify
from openpyxl import load_workbook


# ── Paths / environment ─────────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent

# Load .env/.myenv from the same directory as this script.
# This is important when running under systemd.
try:
    from dotenv import load_dotenv

    load_dotenv(BASE_DIR / ".env")
    load_dotenv(BASE_DIR / ".myenv")
    load_dotenv()
except Exception:
    pass


REPORT_SCRIPT = BASE_DIR / "switch_capacity_report.py"
REPORT_DIR = BASE_DIR / "reports"
REPORT_DIR.mkdir(exist_ok=True)

API_KEY = os.getenv("SWITCH_API_KEY", "change-this-to-a-long-random-string")

TEAMS_WEBHOOK = (
    os.getenv("webhook_switch_capacity")
    or os.getenv("TEAMS_WEBHOOK_URL")
    or os.getenv("webhook_teams")
)

app = Flask(__name__)


# ── Auth / helpers ──────────────────────────────────────────────────────────

def check_api_key():
    return request.headers.get("X-Api-Key") == API_KEY


def bool_value(value):
    return str(value).strip().lower() in ["true", "yes", "1", "on"]


def clean_error_text(text):
    """
    Removes noisy warning lines before showing anything to users.
    Raw stdout/stderr are still returned in the JSON response for troubleshooting.
    """
    if not text:
        return ""

    skip_phrases = [
        "ConflictingConfigurationWarning",
        "Native Python logging configuration has been detected",
        "Nornir logging is enabled too",
        "Please set logging.enabled config to False",
        "https://nornir.readthedocs.io",
        "warnings.warn",
    ]

    cleaned_lines = []
    for line in str(text).splitlines():
        if any(phrase in line for phrase in skip_phrases):
            continue
        cleaned_lines.append(line)

    return "\n".join(cleaned_lines).strip()


def is_no_switch_found_error(text, target):
    text_l = str(text).lower()
    target_l = str(target).lower()

    return (
        "no switches found in nornir inventory" in text_l
        or f"no switches found in nornir inventory for cab {target_l}" in text_l
        or "no switches found" in text_l
    )


# ── Adaptive input card ─────────────────────────────────────────────────────

def input_card():
    """
    Optional endpoint card. Main keyword flow uses the Power Automate card,
    but this is still useful for testing / future reuse.
    """
    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            {
                "type": "TextBlock",
                "text": "Switch Capacity Report",
                "weight": "Bolder",
                "size": "Medium",
                "wrap": True,
            },
            {
                "type": "TextBlock",
                "text": "Enter the building and cab details below.",
                "isSubtle": True,
                "wrap": True,
            },
            {
                "type": "Input.Text",
                "id": "building",
                "label": "Building",
                "placeholder": "Example: 86",
            },
            {
                "type": "Input.Text",
                "id": "cab",
                "label": "Cab number",
                "placeholder": "Example: 102",
            },
            {
                "type": "Input.Text",
                "id": "weeks",
                "label": "Number of weeks unused",
                "value": "1",
                "placeholder": "Example: 4",
            },
            {
                "type": "Input.Toggle",
                "id": "email_report",
                "title": "Email me the Excel report",
                "value": "true",
                "valueOn": "true",
                "valueOff": "false",
            },
            {
                "type": "Input.Text",
                "id": "requester_upn",
                "label": "Your UCL email address",
                "placeholder": "Example: d.ansong@ucl.ac.uk",
            },
        ],
        "actions": [
            {
                "type": "Action.Submit",
                "title": "Run Report",
                "data": {
                    "action": "switch_capacity_report"
                },
            }
        ],
    }


# ── Excel report reader ─────────────────────────────────────────────────────

def read_report_rows(xlsx_path, max_preview=40):
    preview = []
    total = 0

    if not xlsx_path.exists():
        return total, preview

    wb = load_workbook(xlsx_path, data_only=True)

    if "Switch Capacity" not in wb.sheetnames:
        return total, preview

    ws = wb["Switch Capacity"]

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue

        if str(row[0]).strip().lower() == "total free ports":
            break

        total += 1

        if len(preview) < max_preview:
            preview.append(
                {
                    "cab": row[0],
                    "hostname": row[1],
                    "port": row[2],
                    "vlan": row[3],
                    "outlet": row[4],
                    "snmp_location": row[5],
                    "last_used": row[6],
                    "uptime": row[7],
                }
            )

    return total, preview


# ── Email ───────────────────────────────────────────────────────────────────

def send_email_with_attachment(to_addr, subject, body, attachment_path):
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "25"))
    smtp_from = os.getenv("SMTP_FROM", "network-reports@ucl.ac.uk")

    if not smtp_host:
        return "SMTP_HOST not set — email not sent"

    msg = EmailMessage()
    msg["From"] = smtp_from
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment_path.name,
        )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as smtp:
        if bool_value(os.getenv("SMTP_STARTTLS", "false")):
            smtp.starttls()

        smtp_user = os.getenv("SMTP_USER")
        smtp_password = os.getenv("SMTP_PASSWORD")

        if smtp_user and smtp_password:
            smtp.login(smtp_user, smtp_password)

        smtp.send_message(msg)

    return "Email sent"


# ── Cards ───────────────────────────────────────────────────────────────────

def input_error_card(title, message, building="", cab="", weeks="", email_status="not sent"):
    generated_time = datetime.now().strftime("%d %b %Y %H:%M")

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            {
                "type": "TextBlock",
                "text": f"⚠️ {title}",
                "weight": "Bolder",
                "size": "Large",
                "wrap": True,
            },
            {
                "type": "Container",
                "style": "attention",
                "spacing": "Medium",
                "items": [
                    {
                        "type": "TextBlock",
                        "text": message,
                        "wrap": True,
                    }
                ],
            },
            {
                "type": "FactSet",
                "spacing": "Medium",
                "facts": [
                    {"title": "Building", "value": str(building or "not supplied")},
                    {"title": "Cab", "value": str(cab or "not supplied")},
                    {"title": "Unused period", "value": str(weeks or "not supplied")},
                    {"title": "Email status", "value": str(email_status)},
                    {"title": "Generated", "value": generated_time},
                ],
            },
        ],
    }


def no_switch_found_card(building, cab, weeks, email_status="not sent"):
    generated_time = datetime.now().strftime("%d %b %Y %H:%M")

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            {
                "type": "TextBlock",
                "text": "⚠️ Cab not found in inventory",
                "weight": "Bolder",
                "size": "Large",
                "wrap": True,
            },
            {
                "type": "TextBlock",
                "text": f"Building {building} | Cab {cab}",
                "isSubtle": True,
                "spacing": "None",
                "wrap": True,
            },
            {
                "type": "Container",
                "style": "attention",
                "spacing": "Medium",
                "items": [
                    {
                        "type": "TextBlock",
                        "text": (
                            f"No switches were found in the Nornir inventory for "
                            f"Building {building}, Cab {cab}."
                        ),
                        "wrap": True,
                    },
                    {
                        "type": "TextBlock",
                        "text": (
                            "Please check the building and cab number, or confirm that "
                            "the switch inventory has the correct building/cab values."
                        ),
                        "wrap": True,
                    },
                ],
            },
            {
                "type": "FactSet",
                "spacing": "Medium",
                "facts": [
                    {"title": "Building", "value": str(building)},
                    {"title": "Cab", "value": str(cab)},
                    {"title": "Unused period", "value": f"{weeks} week(s)"},
                    {"title": "Email status", "value": str(email_status)},
                    {"title": "Generated", "value": generated_time},
                ],
            },
        ],
    }


def result_card(building, cab, weeks, total, preview_rows, email_status, warnings=None):
    warnings = warnings or []

    if total == 0:
        status_icon = "⚠️"
        status_text = "No free ports found"
        total_color = "Attention"
    elif total <= 5:
        status_icon = "⚠️"
        status_text = "Low spare capacity"
        total_color = "Warning"
    else:
        status_icon = "✅"
        status_text = "Free ports available"
        total_color = "Good"

    generated_time = datetime.now().strftime("%d %b %Y %H:%M")

    body = [
        {
            "type": "TextBlock",
            "text": f"{status_icon} Switch Capacity Report",
            "weight": "Bolder",
            "size": "Large",
            "wrap": True,
        },
        {
            "type": "TextBlock",
            "text": f"Building {building} | Cab {cab}",
            "isSubtle": True,
            "spacing": "None",
            "wrap": True,
        },
        {
            "type": "Container",
            "style": "emphasis",
            "spacing": "Medium",
            "items": [
                {
                    "type": "ColumnSet",
                    "columns": [
                        {
                            "type": "Column",
                            "width": "stretch",
                            "items": [
                                {
                                    "type": "TextBlock",
                                    "text": "Total free ports",
                                    "isSubtle": True,
                                    "wrap": True,
                                },
                                {
                                    "type": "TextBlock",
                                    "text": str(total),
                                    "size": "ExtraLarge",
                                    "weight": "Bolder",
                                    "color": total_color,
                                    "spacing": "None",
                                    "wrap": True,
                                },
                            ],
                        },
                        {
                            "type": "Column",
                            "width": "stretch",
                            "items": [
                                {
                                    "type": "TextBlock",
                                    "text": "Status",
                                    "isSubtle": True,
                                    "wrap": True,
                                },
                                {
                                    "type": "TextBlock",
                                    "text": status_text,
                                    "weight": "Bolder",
                                    "color": total_color,
                                    "spacing": "None",
                                    "wrap": True,
                                },
                            ],
                        },
                    ],
                }
            ],
        },
        {
            "type": "FactSet",
            "spacing": "Medium",
            "facts": [
                {"title": "Building", "value": str(building)},
                {"title": "Cab", "value": str(cab)},
                {"title": "Unused period", "value": f"{weeks} week(s)"},
                {"title": "Email status", "value": str(email_status)},
                {"title": "Generated", "value": generated_time},
            ],
        },
    ]

    if preview_rows:
        grouped = {}

        for r in preview_rows:
            switch_name = str(r.get("hostname") or "Unknown switch").strip()
            grouped.setdefault(switch_name, []).append(r)

        body.append(
            {
                "type": "TextBlock",
                "text": "Free ports by switch",
                "weight": "Bolder",
                "spacing": "Medium",
                "wrap": True,
            }
        )

        for switch_name in sorted(grouped.keys()):
            rows = grouped[switch_name]
            port_lines = []

            for r in rows:
                port = str(r.get("port") or "").strip()
                vlan = str(r.get("vlan") or "").strip()
                outlet = str(r.get("outlet") or "").strip()
                last_used = str(r.get("last_used") or "").strip()
                uptime = str(r.get("uptime") or "").strip()

                line_parts = [
                    f"{port}",
                    f"VLAN {vlan}" if vlan else "VLAN unknown",
                ]

                if outlet:
                    line_parts.append(f"Outlet {outlet}")

                if last_used:
                    line_parts.append(f"Last used {last_used}")

                if uptime:
                    line_parts.append(f"Uptime {uptime}")

                port_lines.append(" | ".join(line_parts))

            body.append(
                {
                    "type": "Container",
                    "separator": True,
                    "spacing": "Medium",
                    "items": [
                        {
                            "type": "TextBlock",
                            "text": f"Switch: {switch_name} — {len(rows)} free port(s)",
                            "weight": "Bolder",
                            "wrap": True,
                        },
                        {
                            "type": "TextBlock",
                            "text": "\n".join(port_lines),
                            "fontType": "Monospace",
                            "wrap": True,
                            "spacing": "Small",
                        },
                    ],
                }
            )

        if total > len(preview_rows):
            body.append(
                {
                    "type": "TextBlock",
                    "text": (
                        f"Showing first {len(preview_rows)} free port(s). "
                        "Full report is available in the Excel email attachment."
                    ),
                    "isSubtle": True,
                    "wrap": True,
                    "spacing": "Medium",
                }
            )

    if warnings:
        body.append(
            {
                "type": "Container",
                "style": "attention",
                "spacing": "Medium",
                "items": [
                    {
                        "type": "TextBlock",
                        "text": "Warnings",
                        "weight": "Bolder",
                        "wrap": True,
                    },
                    {
                        "type": "TextBlock",
                        "text": "\n".join(str(w) for w in warnings[:5]),
                        "wrap": True,
                    },
                ],
            }
        )

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": body,
    }


# ── Teams webhook ───────────────────────────────────────────────────────────

def post_card_to_teams(card):
    if not TEAMS_WEBHOOK:
        return False, "webhook_switch_capacity not set in environment"

    payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": card,
            }
        ],
    }

    try:
        print("[Teams] Posting to webhook…", flush=True)
        response = requests.post(TEAMS_WEBHOOK, json=payload, timeout=30)
        print(f"[Teams] Response: {response.status_code} {response.text!r}", flush=True)

        if response.status_code >= 300:
            return False, f"Teams webhook failed: {response.status_code} {response.text}"

        return True, "Posted to Teams"

    except Exception as exc:
        return False, f"Teams webhook error: {exc}"


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def home():
    return jsonify(
        {
            "status": "ok",
            "message": "Switch Capacity API is running",
            "endpoints": {
                "health": "/api/health",
                "debug": "/api/debug",
                "card": "/api/card",
                "query": "/api/query",
                "test_teams": "/api/test-teams",
            },
        }
    )


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/api/debug", methods=["GET"])
def debug():
    """Quick sanity-check — shows which env vars are loaded without exposing secrets."""
    return jsonify(
        {
            "TEAMS_WEBHOOK_set": bool(TEAMS_WEBHOOK),
            "SMTP_HOST": os.getenv("SMTP_HOST"),
            "SMTP_PORT": os.getenv("SMTP_PORT"),
            "SMTP_FROM": os.getenv("SMTP_FROM"),
            "API_KEY_set": bool(API_KEY),
            "REPORT_SCRIPT_exists": REPORT_SCRIPT.exists(),
            "REPORT_DIR": str(REPORT_DIR),
        }
    )


@app.route("/api/test-teams", methods=["POST"])
def test_teams():
    if not check_api_key():
        return jsonify({"error": "unauthorised"}), 401

    card_json = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            {
                "type": "TextBlock",
                "text": "Switch Capacity API Teams Test",
                "weight": "Bolder",
                "size": "Medium",
                "wrap": True,
            },
            {
                "type": "TextBlock",
                "text": f"Test sent at {datetime.now().strftime('%d %b %Y %H:%M:%S')}",
                "wrap": True,
            },
        ],
    }

    teams_sent, teams_message = post_card_to_teams(card_json)

    return jsonify(
        {
            "teams_sent": teams_sent,
            "teams_message": teams_message,
            "TEAMS_WEBHOOK_set": bool(TEAMS_WEBHOOK),
        }
    )


@app.route("/api/card", methods=["GET"])
def card():
    if not check_api_key():
        return jsonify({"error": "unauthorised"}), 401

    return jsonify({"adaptive_card": input_card()})


@app.route("/api/query", methods=["POST"])
def query():
    if not check_api_key():
        return jsonify({"error": "unauthorised"}), 401

    payload = request.get_json(silent=True) or {}

    building = str(payload.get("building", "")).strip()
    cab = str(payload.get("cab", payload.get("cab_no", ""))).strip()
    weeks_raw = payload.get("weeks", payload.get("unused_weeks", 4))
    email_report = bool_value(payload.get("email_report", False))
    requester_upn = str(payload.get("requester_upn", "")).strip()

    include_admin_down = bool_value(payload.get("include_admin_down", False))
    include_trunks = bool_value(payload.get("include_trunks", False))

    if not building or not cab:
        card_json = input_error_card(
            title="Missing required details",
            message="Building and cab number are required. Please run the report again and enter both values.",
            building=building,
            cab=cab,
            weeks=weeks_raw,
            email_status="not sent",
        )

        teams_sent, teams_message = post_card_to_teams(card_json)

        return jsonify(
            {
                "ok": False,
                "error_type": "missing_required_fields",
                "message": "building and cab are required",
                "adaptive_card": card_json,
                "teams_sent": teams_sent,
                "teams_message": teams_message,
            }
        ), 200

    try:
        weeks = int(str(weeks_raw).strip())
    except Exception:
        card_json = input_error_card(
            title="Invalid unused period",
            message="The number of weeks unused must be a whole number, for example 1, 4, or 8.",
            building=building,
            cab=cab,
            weeks=weeks_raw,
            email_status="not sent",
        )

        teams_sent, teams_message = post_card_to_teams(card_json)

        return jsonify(
            {
                "ok": False,
                "error_type": "invalid_weeks",
                "message": "weeks must be a whole number",
                "adaptive_card": card_json,
                "teams_sent": teams_sent,
                "teams_message": teams_message,
            }
        ), 200

    if weeks < 1:
        weeks = 1

    target = f"{building}/{cab}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outfile = REPORT_DIR / f"capacity_{building}_{cab}_{timestamp}.xlsx"

    cmd = [
        sys.executable,
        str(REPORT_SCRIPT),
        "--target",
        target,
        "--unused-for",
        f"{weeks}w",
        "--xlsx",
        str(outfile),
    ]

    if include_admin_down:
        cmd.append("--include-admin-down")

    if include_trunks:
        cmd.append("--include-trunks")

    print(
        f"[Request] building={building} cab={cab} weeks={weeks} "
        f"email_report={email_report} requester={requester_upn}",
        flush=True,
    )

    proc = subprocess.run(
        cmd,
        cwd=BASE_DIR,
        text=True,
        capture_output=True,
        timeout=1800,
    )

    combined_output = f"{proc.stdout}\n{proc.stderr}"
    cleaned_error = clean_error_text(combined_output)

    if proc.returncode != 0:
        if is_no_switch_found_error(combined_output, target):
            card_json = no_switch_found_card(
                building=building,
                cab=cab,
                weeks=weeks,
                email_status="not sent",
            )

            teams_sent, teams_message = post_card_to_teams(card_json)

            return jsonify(
                {
                    "ok": False,
                    "error_type": "no_switches_found",
                    "message": f"No switches found in inventory for Building {building}, Cab {cab}",
                    "adaptive_card": card_json,
                    "teams_sent": teams_sent,
                    "teams_message": teams_message,
                    "email_status": "not sent",
                    "stdout": proc.stdout,
                    "stderr": proc.stderr,
                }
            ), 200

        friendly_warning = (
            "The report could not be completed. Please check the API logs for details."
        )

        if cleaned_error:
            friendly_warning = cleaned_error[:1000]

        card_json = result_card(
            building=building,
            cab=cab,
            weeks=weeks,
            total=0,
            preview_rows=[],
            email_status="not sent",
            warnings=[friendly_warning],
        )

        teams_sent, teams_message = post_card_to_teams(card_json)

        return jsonify(
            {
                "ok": False,
                "error_type": "report_failed",
                "adaptive_card": card_json,
                "teams_sent": teams_sent,
                "teams_message": teams_message,
                "email_status": "not sent",
                "stdout": proc.stdout,
                "stderr": proc.stderr,
            }
        ), 500

    total, preview_rows = read_report_rows(outfile)

    if total == 0:
        match = re.search(r"Total free ports found:\s*(\d+)", proc.stdout)
        total = int(match.group(1)) if match else 0

    email_status = "not requested"

    if email_report:
        if not requester_upn:
            email_status = "email requested but no email address supplied"
        else:
            try:
                email_status = send_email_with_attachment(
                    to_addr=requester_upn,
                    subject=f"Switch Capacity Report {target}",
                    body=(
                        f"Please find attached the switch capacity report for CAB {target}.\n\n"
                        f"Requested unused period: {weeks} week(s)\n"
                        f"Total free ports: {total}\n"
                    ),
                    attachment_path=outfile,
                )
            except Exception as exc:
                email_status = f"Email error: {exc}"

    card_json = result_card(
        building=building,
        cab=cab,
        weeks=weeks,
        total=total,
        preview_rows=preview_rows,
        email_status=email_status,
    )

    teams_sent, teams_message = post_card_to_teams(card_json)

    return jsonify(
        {
            "ok": True,
            "adaptive_card": card_json,
            "teams_sent": teams_sent,
            "teams_message": teams_message,
            "email_status": email_status,
            "total_free_ports": total,
            "xlsx_path": str(outfile),
            "stdout": proc.stdout,
        }
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8092)