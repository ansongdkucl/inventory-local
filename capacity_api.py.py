#!/usr/bin/env python3

import os
import re
import sys
import smtplib
import subprocess
from datetime import datetime
from pathlib import Path
from email.message import EmailMessage

from flask import Flask, request, jsonify
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass


BASE_DIR = Path(__file__).resolve().parent
REPORT_SCRIPT = BASE_DIR / "switch_capacity_report.py"
REPORT_DIR = BASE_DIR / "reports"
REPORT_DIR.mkdir(exist_ok=True)

API_KEY = os.getenv("SWITCH_API_KEY", "change-this-to-a-long-random-string")

app = Flask(__name__)


def check_api_key():
    supplied = request.headers.get("X-Api-Key")
    return supplied == API_KEY


def bool_value(value):
    return str(value).lower() in ["true", "yes", "1", "on"]


def input_card():
    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            {
                "type": "TextBlock",
                "text": "Switch Capacity Report",
                "weight": "Bolder",
                "size": "Medium",
                "wrap": True,
            },
            {
                "type": "Input.Text",
                "id": "building",
                "label": "Building",
                "placeholder": "Example: 86",
            },
            {
                "type": "Input.Text",
                "id": "cab",
                "label": "Cab No",
                "placeholder": "Example: 100",
            },
            {
                "type": "Input.Number",
                "id": "weeks",
                "label": "Number of weeks unused",
                "value": 4,
                "min": 1,
                "max": 104,
            },
            {
                "type": "Input.Toggle",
                "id": "email_report",
                "title": "Send Excel report by email",
                "value": "false",
                "valueOn": "true",
                "valueOff": "false",
            },
            {
                "type": "Input.Text",
                "id": "requester_upn",
                "label": "Email address",
                "placeholder": "Example: cceadan@ucl.ac.uk",
            },
        ],
        "actions": [
            {
                "type": "Action.Submit",
                "title": "Run Report",
                "data": {
                    "action": "switch_capacity_report"
                },
            }
        ],
    }


def read_report_preview(xlsx_path, max_rows=10):
    rows = []

    if not xlsx_path.exists():
        return rows

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Switch Capacity"]

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue

        if str(row[0]).strip().lower() == "total free ports":
            break

        rows.append(
            {
                "cab": row[0],
                "hostname": row[1],
                "port": row[2],
                "vlan": row[3],
                "outlet": row[4],
                "snmp_location": row[5],
                "last_used": row[6],
                "uptime": row[7],
            }
        )

    return rows[:max_rows]


def send_email_with_attachment(to_addr, subject, body, attachment_path):
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "25"))
    smtp_from = os.getenv("SMTP_FROM", "network-reports@ucl.ac.uk")

    if not smtp_host:
        return False, "SMTP_HOST is not set, so email was not sent."

    msg = EmailMessage()
    msg["From"] = smtp_from
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment_path.name,
        )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as smtp:
        if bool_value(os.getenv("SMTP_STARTTLS", "false")):
            smtp.starttls()

        smtp_user = os.getenv("SMTP_USER")
        smtp_password = os.getenv("SMTP_PASSWORD")

        if smtp_user and smtp_password:
            smtp.login(smtp_user, smtp_password)

        smtp.send_message(msg)

    return True, "Email sent."


def result_card(building, cab, weeks, total, preview_rows, email_status, warnings=None):
    warnings = warnings or []

    facts = [
        {"title": "Building", "value": str(building)},
        {"title": "Cab", "value": str(cab)},
        {"title": "Requested unused period", "value": f"{weeks} week(s)"},
        {"title": "Total free ports", "value": str(total)},
        {"title": "Email status", "value": email_status},
        {"title": "Generated", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]

    body = [
        {
            "type": "TextBlock",
            "text": f"Switch Capacity Report - Building {building} Cab {cab}",
            "weight": "Bolder",
            "size": "Medium",
            "wrap": True,
        },
        {
            "type": "FactSet",
            "facts": facts,
        },
    ]

    if preview_rows:
        lines = []
        for r in preview_rows:
            lines.append(
                f"{r['hostname']} | {r['port']} | VLAN {r['vlan']} | {r['last_used']} | uptime {r['uptime']}"
            )

        body.append(
            {
                "type": "TextBlock",
                "text": "Top results",
                "weight": "Bolder",
                "wrap": True,
                "spacing": "Medium",
            }
        )

        body.append(
            {
                "type": "TextBlock",
                "text": "\n".join(lines),
                "wrap": True,
                "fontType": "Monospace",
            }
        )

    if warnings:
        body.append(
            {
                "type": "TextBlock",
                "text": "Warnings",
                "weight": "Bolder",
                "color": "Warning",
                "wrap": True,
            }
        )
        body.append(
            {
                "type": "TextBlock",
                "text": "\n".join(warnings[:5]),
                "wrap": True,
            }
        )

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": body,
    }


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/api/card", methods=["GET"])
def get_card():
    if not check_api_key():
        return jsonify({"error": "unauthorised"}), 401

    return jsonify({"adaptive_card": input_card()})


@app.route("/api/query", methods=["POST"])
def run_query():
    if not check_api_key():
        return jsonify({"error": "unauthorised"}), 401

    payload = request.get_json(force=True)

    building = str(payload.get("building", "")).strip()
    cab = str(payload.get("cab", payload.get("cab_no", ""))).strip()
    weeks = int(payload.get("weeks", payload.get("unused_weeks", 4)))
    email_report = bool_value(payload.get("email_report", False))
    requester_upn = str(payload.get("requester_upn", "")).strip()

    include_admin_down = bool_value(payload.get("include_admin_down", False))
    include_trunks = bool_value(payload.get("include_trunks", False))

    if not building or not cab:
        return jsonify({"error": "building and cab are required"}), 400

    target = f"{building}/{cab}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outfile = REPORT_DIR / f"capacity_{building}_{cab}_{timestamp}.xlsx"

    cmd = [
        sys.executable,
        str(REPORT_SCRIPT),
        "--target",
        target,
        "--unused-for",
        f"{weeks}w",
        "--xlsx",
        str(outfile),
    ]

    if include_admin_down:
        cmd.append("--include-admin-down")

    if include_trunks:
        cmd.append("--include-trunks")

    proc = subprocess.run(
        cmd,
        cwd=BASE_DIR,
        text=True,
        capture_output=True,
        timeout=1800,
    )

    if proc.returncode != 0:
        card = result_card(
            building,
            cab,
            weeks,
            total=0,
            preview_rows=[],
            email_status="not sent",
            warnings=[proc.stderr.strip() or proc.stdout.strip()],
        )

        return jsonify(
            {
                "ok": False,
                "adaptive_card": card,
                "stdout": proc.stdout,
                "stderr": proc.stderr,
            }
        ), 500

    total_match = re.search(r"Total free ports found:\s*(\d+)", proc.stdout)
    total = int(total_match.group(1)) if total_match else 0

    preview_rows = read_report_preview(outfile)

    email_status = "not requested"

    if email_report:
        if not requester_upn:
            email_status = "email requested but no email address supplied"
        else:
            sent, message = send_email_with_attachment(
                to_addr=requester_upn,
                subject=f"Switch Capacity Report {target}",
                body=(
                    f"Please find attached the switch capacity report for CAB {target}.\n\n"
                    f"Requested unused period: {weeks} week(s)\n"
                    f"Total free ports: {total}\n"
                ),
                attachment_path=outfile,
            )
            email_status = message

    card = result_card(
        building=building,
        cab=cab,
        weeks=weeks,
        total=total,
        preview_rows=preview_rows,
        email_status=email_status,
    )

    return jsonify(
        {
            "ok": True,
            "adaptive_card": card,
            "total_free_ports": total,
            "xlsx_path": str(outfile),
            "stdout": proc.stdout,
        }
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8088)