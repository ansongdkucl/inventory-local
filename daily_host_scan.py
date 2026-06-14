from netmiko import ConnectHandler
import netmiko
import json
import nmap
import os
from pathlib import Path
import re
from copy import copy
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import smtplib
import requests
from requests.auth import HTTPBasicAuth
import urllib3
from openpyxl import load_workbook
from inventory import detect_platform, get_switch_details

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

LOCAL_EXCEL_FILE = Path("inventory.xlsx").resolve()
SHEET_NAME = "inventory"
LEGACY_SHEET_NAME = "Sheet1"
IP_COLUMN = "IP Address"
HOSTNAME_COLUMN = "hostname"
SWITCH_NUM_COLUMN = "switch_num"
MODEL_COLUMN = "model"
PORTS_COLUMN = "ports"
VENDOR_COLUMN = "vendor"
LOCATION_COLUMN = "location"
MISINVENTORY_SHEET = "misinventory"
MISOLARWINDS_SHEET = "misolarwinds"
#SOLARWINDS_URL = os.environ.get("SOLARWINDS_URL")

SOLARWINDS_URL = (
    "https://service-health.isd.ucl.ac.uk:17778/"
    "SolarWinds/InformationService/v3/Json/Query"
)
SOLARWINDS_USERNAME = os.environ.get("SOLARWINDS_USERNAME") or os.environ.get("username")
SOLARWINDS_PASSWORD = os.environ.get("SOLARWINDS_PASSWORD") or os.environ.get("passwordAD")

# Ignore list - devices to exclude from discovery
ig_list = [
    '172.17.49.240','172.22.29.200','172.17.107.250','172.17.107.251',
    '172.17.57.240','172.17.57.241','172.17.57.242','172.17.57.243',
    '172.17.57.244','172.17.57.249'
    # Add other IPs to ignore here
]

ignored_subnet_prefixes = (
    '172.17.87.',
    '172.17.96.',
    '172.17.92.',
    '172.17.112.',
    '172.17.115.',
    '172.17.116.',
    '172.17.125.'
)

# List Variables
host_email = []
new_list = []
host_m = []
all_live_hosts = []  # To store all live hosts for output file
filtered_values = []  # Initialize filtered_values list
discovered_devices = {}  # Dictionary to store unique devices with their details

# Email Variables
fromaddr = "cceadan@ucl.ac.uk"
toaddr = "cceadan@ucl.ac.uk"
server = None

# Environmental Variables
user = os.environ['username']
secret = os.environ['secret']
password = os.environ.get('password')
aruba_pw = os.environ.get('passwordAD')
aruba_secret = os.environ.get('aruba_secret')  # New variable for 172.23/172.30 subnets


# Scan multiple subnets to discover all live hosts on the network.
# Returns the total number of connected ip addresses.
def live_scan():
    print('Scanning subnets to discover all live hosts on the network.')
    nm = nmap.PortScanner()
    # Added 172.22.x.x subnet to the scan
    hosts = '172.17.0-130.1-253 172.16.1.0/24 172.23.1-17.1-253 172.30.1-2.1-253 172.22.0-30.1-253'
    n = nm.scan(hosts=hosts, arguments='-n -sP -PE')
    live_hosts = nm.all_hosts()
    len_hosts = len(live_hosts)
    print('There are {} live hosts on the specified subnets'.format(len_hosts))
    
    # Save all live hosts to the global variable for output file
    global all_live_hosts
    all_live_hosts = live_hosts.copy()
    
    return live_hosts


# Remove selected devices from list
# Host ending with (252|254|253) tend to be router and gateway addresses
def host_routers(live_hosts):
    filtered_values = list(filter(lambda v: re.match('172.17.\d+.(112|115|252|254|253)', v), live_hosts))
    filtered_values.extend(
        host for host in live_hosts
        if host.startswith(ignored_subnet_prefixes)
    )
    filter_hosts = len(filtered_values)
    print('There are {} Excluded hosts:'.format(filter_hosts))
    
    for x in ig_list:
        filtered_values.append(x) 
         
    for x in filtered_values: 
        if x in live_hosts: 
            live_hosts.remove(x) 
    return live_hosts 
 

def sort_ips(ips):
    def sort_key(ip):
        return tuple(int(part) for part in ip.split("."))

    return sorted(ips, key=sort_key)


def get_headers(ws):
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value

        if value:
            headers[str(value).strip().lower()] = col

    return headers


def get_inventory_ips(ws, ip_col):
    ips = set()

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=ip_col).value

        if value:
            ips.add(str(value).strip())

    return {ip for ip in ips if ip}


def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    return wb.create_sheet(sheet_name)


def get_inventory_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]

    if LEGACY_SHEET_NAME in wb.sheetnames:
        return wb[LEGACY_SHEET_NAME]

    raise RuntimeError(
        f"Inventory workbook must contain '{SHEET_NAME}' worksheet "
        f"or legacy '{LEGACY_SHEET_NAME}' worksheet."
    )


def get_next_available_row(ws):
    row = ws.max_row

    while row > 1:
        has_value = any(
            ws.cell(row=row, column=col).value not in (None, "")
            for col in range(1, ws.max_column + 1)
        )

        if has_value:
            break

        row -= 1

    return row + 1


def copy_previous_row_style(ws, target_row):
    source_row = target_row - 1

    if source_row < 2:
        return

    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format


def vendor_for_ip(ip):
    platform_type = detect_platform(ip)

    if platform_type == "aruba_cx":
        return "aruba_aoscx"

    if platform_type == "aruba_os":
        return "aruba_os"

    return "cisco_os"


def append_inventory_device(ws, headers, ip, details):
    row = get_next_available_row(ws)
    copy_previous_row_style(ws, row)

    values = {
        IP_COLUMN.lower(): ip,
        HOSTNAME_COLUMN.lower(): details.get("hostname"),
        SWITCH_NUM_COLUMN.lower(): details.get("switch_num"),
        MODEL_COLUMN.lower(): details.get("model"),
        PORTS_COLUMN.lower(): details.get("ports"),
        VENDOR_COLUMN.lower(): vendor_for_ip(ip),
        LOCATION_COLUMN.lower(): details.get("snmp_location"),
    }

    for column_name, value in values.items():
        if column_name in headers and value is not None:
            ws.cell(row=row, column=headers[column_name]).value = value

    return row


def add_missing_devices_to_inventory(ws, headers, missing_inventory):
    required_columns = [
        IP_COLUMN,
        HOSTNAME_COLUMN,
        SWITCH_NUM_COLUMN,
        MODEL_COLUMN,
        PORTS_COLUMN,
        VENDOR_COLUMN,
        LOCATION_COLUMN,
    ]
    missing_columns = [
        column for column in required_columns
        if column.lower() not in headers
    ]

    if missing_columns:
        raise RuntimeError(f"Missing required columns in inventory workbook: {missing_columns}")

    added_devices = {}
    unresolved_ips = set()

    for ip in sort_ips(missing_inventory):
        print(f"[*] Adding missing inventory device {ip}...", end=" ")
        details = get_switch_details(ip)

        if not details:
            print("failed")
            unresolved_ips.add(ip)
            continue

        row = append_inventory_device(ws, headers, ip, details)
        added_devices[ip] = {
            **details,
            "row": row,
            "vendor": vendor_for_ip(ip),
        }
        print(f"added to row {row}")

    return added_devices, unresolved_ips


def write_ip_sheet(ws, ips):
    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1).value = IP_COLUMN

    for index, ip in enumerate(sort_ips(ips), start=2):
        ws.cell(row=index, column=1).value = ip


def update_missing_ip_sheets(live_hosts, solarwinds_ips):
    wb = load_workbook(LOCAL_EXCEL_FILE)
    inventory_ws = get_inventory_sheet(wb)
    headers = get_headers(inventory_ws)

    if IP_COLUMN.lower() not in headers:
        raise RuntimeError(f"Missing required column in inventory workbook: {IP_COLUMN}")

    ip_col = headers[IP_COLUMN.lower()]
    inventory_ips = get_inventory_ips(inventory_ws, ip_col)
    live_ip_set = {str(ip).strip() for ip in live_hosts if str(ip).strip()}
    missing_inventory = live_ip_set - inventory_ips
    missing_solarwinds = live_ip_set - solarwinds_ips
    added_inventory, unresolved_inventory = add_missing_devices_to_inventory(
        inventory_ws,
        headers,
        missing_inventory,
    )

    write_ip_sheet(get_or_create_sheet(wb, MISINVENTORY_SHEET), unresolved_inventory)
    write_ip_sheet(get_or_create_sheet(wb, MISOLARWINDS_SHEET), missing_solarwinds)

    wb.save(LOCAL_EXCEL_FILE)
    wb.close()

    print(f"[+] Inventory IPs: {len(inventory_ips)}")
    print(f"[+] Missing inventory devices added to {SHEET_NAME} worksheet: {len(added_inventory)}")
    print(f"[+] Missing from inventory written to {MISINVENTORY_SHEET} worksheet: {len(unresolved_inventory)}")
    print(f"[+] Missing from SolarWinds written to {MISOLARWINDS_SHEET} worksheet: {len(missing_solarwinds)}")

    return unresolved_inventory, missing_solarwinds, added_inventory


def get_solarwinds_ips():
    if not SOLARWINDS_URL:
        raise RuntimeError("Missing environment variable: SOLARWINDS_URL")

    if not SOLARWINDS_USERNAME:
        raise RuntimeError("Missing SolarWinds username. Set SOLARWINDS_USERNAME or username.")

    if not SOLARWINDS_PASSWORD:
        raise RuntimeError("Missing SolarWinds password. Set SOLARWINDS_PASSWORD or password.")

    print("[*] Querying SolarWinds using HTTPBasicAuth...")

    swql = """
    SELECT n.IP_Address
    FROM Orion.Nodes n
    WHERE n.CustomProperties.Team LIKE '%Tech%'
    """

    max_retries = 3
    response = None

    for attempt in range(max_retries):
        try:
            response = requests.post(
                SOLARWINDS_URL,
                auth=HTTPBasicAuth(SOLARWINDS_USERNAME, SOLARWINDS_PASSWORD),
                verify=False,
                headers={"Content-Type": "application/json"},
                data=json.dumps({"query": swql}),
                timeout=30,
            )
            break
        except requests.exceptions.ConnectionError as e:
            if attempt < max_retries - 1:
                print(f"[!] Connection error (attempt {attempt + 1}/{max_retries}): {e}")
                time.sleep(2)
            else:
                print("[!] Failed to connect to SolarWinds server after multiple attempts.")
                print("    This may be due to network restrictions (e.g. firewall, VPN required).")
                print(f"    Server URL: {SOLARWINDS_URL}")
                print("    Ensure you are connected to the appropriate network.")
                raise

    if response is None:
        raise RuntimeError("SolarWinds query did not return a response.")

    if response.status_code != 200:
        raise Exception(f"SWIS Query failed: {response.status_code} - {response.text}")

    data = response.json()
    ips = {row["IP_Address"] for row in data.get("results", []) if row.get("IP_Address")}

    print(f"[+] Devices in SolarWinds TEAM=Tech: {len(ips)}")
    return ips




# Save all live hosts to an output file
def save_live_hosts_to_file(live_hosts):
    output_file = "found_devices.txt"
    try:
        with open(output_file, 'w') as f:
            f.write("Live Hosts Found During Scan:\n")
            f.write("=" * 50 + "\n")
            f.write(f"Scan Timestamp: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 50 + "\n\n")
            
            # Define data center subnets
            data_center_subnets = ['172.17.112.', '172.17.116.', '172.17.87.']
            
            # Separate data center hosts from other hosts
            data_center_hosts = []
            other_hosts = []
            
            for host in sorted(live_hosts):
                if any(host.startswith(subnet) for subnet in data_center_subnets):
                    data_center_hosts.append(host)
                else:
                    other_hosts.append(host)
            
            # Write data center hosts with heading
            if data_center_hosts:
                f.write("Data Centre:\n")
                f.write("-" * 30 + "\n")
                for host in sorted(data_center_hosts):
                    f.write(f"{host}\n")
                f.write("\n")
            
            # Write other hosts without subnet headings, just as a flat list
            if other_hosts:
                f.write("Other Hosts:\n")
                f.write("-" * 30 + "\n")
                for host in sorted(other_hosts):
                    f.write(f"{host}\n")
        
        print(f"All live hosts saved to {output_file}")
    except Exception as e:
        print(f"Error saving live hosts to file: {e}")


def get_device_location(connect, ip):
    """
    Get device location information from the device.
    """
    try:
        # For Cisco devices
        if not ip.startswith(('172.22.', '172.23.', '172.30.')):
            output = connect.send_command('show snmp location')
            return output.strip() if output.strip() else "Location not configured"
        else:
            # For Aruba devices - try different commands
            try:
                output = connect.send_command('show snmp')
                for line in output.splitlines():
                    if 'location' in line.lower():
                        location = line.split(':', 1)[-1].strip()
                        if location:
                            return location
            except:
                pass
            
            # Alternative command for Aruba
            try:
                output = connect.send_command('show system')
                for line in output.splitlines():
                    if 'location' in line.lower():
                        location = line.split(':', 1)[-1].strip()
                        if location:
                            return location
            except:
                pass
                
            return "Location not configured"
    except Exception as e:
        return f"Error retrieving location: {str(e)}"


def login(host_m):
    global discovered_devices
    discovered_devices = {}  # Reset the dictionary
    
    print('\n' + '='*80)
    print('Logging into Newly Discovered Devices')
    print('='*80 + '\n')
    
    for ip in host_m:
        print(f"Connecting to {ip}...", end=' ')
        
        # Determine device type and credentials based on IP subnet
        if ip.startswith('172.22.'):
            device_type = 'aruba_aoscx'
            login_password = aruba_pw
            use_secret = None
            
        elif ip.startswith(('172.23.', '172.30.')):
            device_type = 'aruba_os'
            login_password = password
            use_secret = aruba_secret
            
        else:
            device_type = 'cisco_ios'
            login_password = password
            use_secret = secret
        
        try:
            # Create connection parameters
            device_params = {
                'device_type': device_type,
                'ip': ip,
                'username': user,
                'password': login_password,
                'session_log': 'net.log'
            }
            
            # Add secret if specified
            if use_secret:
                device_params['secret'] = use_secret
            
            connect = ConnectHandler(**device_params)
            
            hostname = connect.find_prompt().strip().replace('#', '').replace('>', '')
            location = get_device_location(connect, ip)
            
            # Store in dictionary to avoid duplicates
            discovered_devices[ip] = {
                'hostname': hostname,
                'location': location
            }
            
            print('✓ Success')
            connect.disconnect()
            
        except netmiko.NetmikoTimeoutException:
            print('✗ Timeout')
            discovered_devices[ip] = {
                'hostname': 'Unknown',
                'location': 'Connection timeout'
            }
        except netmiko.NetmikoAuthenticationException:
            print('✗ Auth Failed')
            discovered_devices[ip] = {
                'hostname': 'Unknown',
                'location': 'Authentication failed'
            }
        except Exception as e:
            print(f'✗ Error: {str(e)[:50]}')
            discovered_devices[ip] = {
                'hostname': 'Unknown',
                'location': f'Error: {str(e)}'
            }
    
    # Print results in formatted style
    print('\n' + '='*80)
    print('DISCOVERED DEVICES NOT IN INVENTORY')
    print('='*80 + '\n')
    
    host_email_list = []
    for ip in sorted(discovered_devices.keys()):
        device = discovered_devices[ip]
        hostname = device['hostname']
        location = device['location']
        
        # Format the output line
        formatted_line = f"{hostname:<30} {ip:<18} <----------------> Location: {location}"
        print(formatted_line)
        host_email_list.append(formatted_line)
    
    print('\n' + '='*80 + '\n')
    
    return host_email_list


def send_email(host_email, subject='New Network Device Alert', heading='DEVICES ON NETWORK BUT NOT IN INVENTORY'):
    print('Sending email notification...')
    time.sleep(2)
    
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = subject

    body = f'{heading}\n'
    body += '=' * 80 + '\n'
    body += f"Scan Time: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
    body += '=' * 80 + '\n\n'
    body += "\n".join(host_email)
    body += f"\n\n{'=' * 80}\n"
    body += f"Total devices found: {len(host_email)}\n"

    msg.attach(MIMEText(body, 'plain'))
    text = msg.as_string()
    
    server = None
    try:
        server = smtplib.SMTP('smtp-server.ucl.ac.uk', 587)
        server.starttls()
        server.sendmail(fromaddr, toaddr, text)
        print("✓ Email sent successfully")
    except Exception as e:
        print(f"✗ Failed to send email: {e}")
    finally:
        if server:
            try:
                server.quit()
            except:
                pass


def main():
    live_hosts = live_scan()
    
    # Save all live hosts to output file
    save_live_hosts_to_file(all_live_hosts)
    
    filtered_live_hosts = host_routers(live_hosts)
    solarwinds_ips = get_solarwinds_ips()
    missing_inventory, missing_solarwinds, added_inventory = update_missing_ip_sheets(
        filtered_live_hosts,
        solarwinds_ips,
    )

    global host_m
    host_m = sort_ips(missing_inventory)

    if len(added_inventory) != 0:
        print(f'\n✓ Added {len(added_inventory)} devices to inventory\n')
        added_email = []

        for ip in sort_ips(added_inventory.keys()):
            device = added_inventory[ip]
            added_email.append(
                f"{device.get('hostname', '-'):<30} "
                f"{ip:<18} "
                f"row={device.get('row', '-')} "
                f"vendor={device.get('vendor', '-')} "
                f"model={device.get('model', '-')} "
                f"ports={device.get('ports', '-')} "
                f"switch_num={device.get('switch_num', '-')} "
                f"location={device.get('snmp_location', '-')}"
            )

        send_email(
            added_email,
            subject='New Network Device Added to Inventory',
            heading='NEW DEVICES DISCOVERED AND ADDED TO INVENTORY',
        )

    if len(host_m) != 0:
        print(f'\n⚠ Found {len(host_m)} devices on the network but could not add them to inventory\n')
        host_email = [
            f"{ip:<18} <----------------> Could not collect details or add to inventory"
            for ip in host_m
        ]
        send_email(host_email)
    else:
        print('\n✓ No devices missing from inventory\n')

    if len(missing_solarwinds) != 0:
        print(f'⚠ Found {len(missing_solarwinds)} devices on the network but not in SolarWinds')
    else:
        print('✓ No devices missing from SolarWinds')


if __name__ == '__main__':
    main()
