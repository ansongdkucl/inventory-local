#!/usr/bin/env python3

# python inventory.py --ping-only
# python inventory.py --ping-only --ping-timeout 1
# python inventory.py --start-row 3000
# python inventory.py --start-row 3000 --no-copy-back
#
# Safer Excel workflow:
# 1. Work on local ./inventory.xlsx
# 2. Save edits locally
# 3. Try to copy final file back to OneDrive
# 4. If copy-back fails, keep the edited local inventory.xlsx
#
# Platform logic:
# 172.22.x.x     = Aruba CX, uses password_AD or passwordAD
# 172.23.x.x     = Aruba OS / S2500, uses aruba_pw
# 172.30.30.x    = Aruba OS / S2500, uses aruba_pw
# Everything else = Cisco IOS, uses password

import os
import argparse
import logging
import subprocess
import shutil
import platform
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv
from netmiko import ConnectHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================
# RUN ID
# ============================================================
RUN_ID = datetime.now().strftime("%Y%m%d-%H%M%S")

# ============================================================
# LOAD ENV
# ============================================================
load_dotenv()

SWITCH_USERNAME = os.getenv("username")

# Cisco login
SWITCH_PASSWORD = os.getenv("password")

# Aruba CX login
SWITCH_PASSWORD_AD = os.getenv("password_AD") or os.getenv("passwordAD")

# Cisco enable
SWITCH_SECRET = os.getenv("secret")

# Aruba OS / S2500 / older Aruba switches
# Used for 172.23.x.x and 172.30.30.x
SWITCH_PASSWORD_ARUBA_OS = os.getenv("aruba_pw")

# Aruba OS enable secret fallback
SWITCH_SECRET_ARUBA_OS = (
    os.getenv("aruba_secret")
    or os.getenv("secret_AOS")
    or os.getenv("secret")
    or SWITCH_PASSWORD_ARUBA_OS
)

# ============================================================
# FILE PATHS
# ============================================================
LOCAL_EXCEL_FILE = Path("inventory.xlsx").resolve()

ONEDRIVE_EXCEL_FILE = Path(
    "/mnt/c/Users/anson/University College London/"
    "ISD.ITSD.CO.Technical Specialists - configs/inventory.xlsx"
)

SHEET_NAME = "Sheet1"

IP_COLUMN = "IP Address"
HOSTNAME_COLUMN = "hostname"
LOCATION_COLUMN = "location"
SWITCH_NUM_COLUMN = "switch_num"
MODEL_COLUMN = "model"
PORTS_COLUMN = "ports"

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

NETWORK_LOG = os.path.join(LOG_DIR, f"network_debug_{RUN_ID}.log")
PING_FAIL_LOG = os.path.join(LOG_DIR, f"ping_failures_{RUN_ID}.txt")

logging.basicConfig(
    filename=NETWORK_LOG,
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

console = logging.StreamHandler()
console.setLevel(logging.INFO)
console.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
logging.getLogger().addHandler(console)

logging.info("=== Script started ===")

RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# ============================================================
# ENV CHECK
# ============================================================
def log_env_status():
    checks = {
        "username": SWITCH_USERNAME,
        "password": SWITCH_PASSWORD,
        "password_AD/passwordAD": SWITCH_PASSWORD_AD,
        "secret": SWITCH_SECRET,
        "aruba_pw": SWITCH_PASSWORD_ARUBA_OS,
        "aruba_secret fallback": SWITCH_SECRET_ARUBA_OS,
    }

    for name, value in checks.items():
        if value:
            logging.info(f"[ENV OK] {name} is set")
        else:
            logging.warning(f"[ENV MISSING] {name} is not set")


def require_credential(value, name, ip):
    if not value:
        logging.error(f"[SKIP] {ip} missing credential: {name}")
        return False
    return True

# ============================================================
# EXCEL COPY / SYNC HELPERS
# ============================================================
def onedrive_lock_file_exists():
    """
    Checks for the temporary Office lock file created when Excel has the workbook open.
    Example: ~$inventory.xlsx
    """
    lock_file = ONEDRIVE_EXCEL_FILE.parent / f"~${ONEDRIVE_EXCEL_FILE.name}"
    return lock_file.exists(), lock_file


def prepare_local_excel_file():
    """
    Use local ./inventory.xlsx.

    If it does not exist, try to copy it from OneDrive once.
    After this, all edits happen locally.
    """
    if LOCAL_EXCEL_FILE.exists():
        logging.info(f"Using local Excel file: {LOCAL_EXCEL_FILE}")
        return

    logging.warning(f"Local inventory.xlsx not found: {LOCAL_EXCEL_FILE}")
    logging.info("Trying to copy inventory.xlsx from OneDrive...")

    if not ONEDRIVE_EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Neither local nor OneDrive Excel file exists.\n"
            f"Local expected here: {LOCAL_EXCEL_FILE}\n"
            f"OneDrive expected here: {ONEDRIVE_EXCEL_FILE}"
        )

    shutil.copy2(ONEDRIVE_EXCEL_FILE, LOCAL_EXCEL_FILE)
    logging.info(f"Copied OneDrive file to local working file: {LOCAL_EXCEL_FILE}")


def copy_result_back_to_onedrive():
    """
    Try to copy the completed local workbook back to OneDrive.

    If this fails, do not crash the script.
    The edited copy remains safely at ./inventory.xlsx.
    """
    try:
        locked, lock_file = onedrive_lock_file_exists()

        if locked:
            logging.warning(
                f"Could not copy back to OneDrive because Excel appears to have the file open.\n"
                f"Lock file found: {lock_file}\n"
                f"Your edits are still saved locally here: {LOCAL_EXCEL_FILE}"
            )
            return False

        shutil.copy2(LOCAL_EXCEL_FILE, ONEDRIVE_EXCEL_FILE)

        logging.info(f"Copied updated workbook back to OneDrive: {ONEDRIVE_EXCEL_FILE}")
        return True

    except Exception as e:
        logging.warning(
            f"Could not copy updated workbook back to OneDrive.\n"
            f"Reason: {e}\n"
            f"Your edits are still saved locally here: {LOCAL_EXCEL_FILE}"
        )
        return False

# ============================================================
# EXCEL HELPERS
# ============================================================
def load_excel():
    prepare_local_excel_file()
    wb = load_workbook(LOCAL_EXCEL_FILE)
    ws = wb[SHEET_NAME]
    return wb, ws


def get_headers(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col
    return headers


def update_cell(ws, row, col, value):
    ws.cell(row=row, column=col).value = value


def highlight_cell(ws, row, col):
    ws.cell(row=row, column=col).fill = RED_FILL


def clean_console_value(value):
    """
    Keeps console output readable by turning None/blank/multiline output
    into a short single-line value.
    """
    if value is None or value == "":
        return "-"

    return " ".join(str(value).split())


def log_row_result(row_num, ip, status, data=None, reason=None):
    """
    Prints one clear result line per Excel row.

    Because this uses logging, the same line appears:
    - on screen
    - in logs/network_debug_<RUN_ID>.log
    """
    data = data or {}

    parts = [
        f"[ROW {row_num}]",
        ip,
        status,
    ]

    if data:
        parts.extend(
            [
                f"hostname={clean_console_value(data.get('hostname'))}",
                f"location={clean_console_value(data.get('snmp_location'))}",
                f"switch_num={clean_console_value(data.get('switch_num'))}",
                f"model={clean_console_value(data.get('model'))}",
                f"ports={clean_console_value(data.get('ports'))}",
            ]
        )

    if reason:
        parts.append(f"reason={clean_console_value(reason)}")

    message = " | ".join(parts)

    if "FAIL" in status or "SKIP" in status:
        logging.warning(message)
    else:
        logging.info(message)

# ============================================================
# PING
# ============================================================
def ping_host(ip, timeout=2):
    try:
        if platform.system().lower() == "windows":
            cmd = ["ping", "-n", "1", "-w", str(timeout * 1000), ip]
        else:
            cmd = ["ping", "-c", "1", "-W", str(timeout), ip]

        return subprocess.run(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        ).returncode == 0

    except Exception:
        return False

# ============================================================
# PLATFORM DETECTION
# ============================================================
def detect_platform(ip):
    """
    Platform selection:

    - 172.22.x.x     = Aruba CX
    - 172.23.x.x     = Aruba OS / S2500 / older Aruba
    - 172.30.30.x    = Aruba OS / S2500 / older Aruba
    - everything else = Cisco IOS
    """
    if ip.startswith("172.22."):
        return "aruba_cx"

    if ip.startswith("172.23."):
        return "aruba_os"

    if ip.startswith("172.30.30."):
        return "aruba_os"

    return "cisco_ios"

# ============================================================
# SAFE COMMAND HELPER
# ============================================================
def send_command_safe(nc, command, read_timeout=20):
    try:
        return nc.send_command(
            command,
            read_timeout=read_timeout,
            expect_string=r"[>#]",
        )
    except Exception as e:
        logging.debug(f"Command failed: {command} | {e}")
        return ""

# ============================================================
# CISCO HELPERS
# ============================================================
def get_cisco_model_from_inventory(nc):
    inv = nc.send_command("show inventory", read_timeout=30)

    for line in inv.splitlines():
        if line.strip().startswith("PID:"):
            return line.split("PID:")[1].split(",")[0].strip()

    return None


def count_cisco_physical_ports(nc):
    iface = nc.send_command("show interfaces status", read_timeout=30)

    physical = (
        "Gi",
        "Fa",
        "Te",
        "Ten",
        "Twe",
        "Fo",
        "Forty",
        "Hu",
        "Hundred",
        "Eth",
    )

    exclude = (
        "Po",
        "Port",
        "Vl",
        "Vlan",
        "Lo",
        "Loop",
        "Ap",
        "Nu",
        "Tu",
        "BDI",
    )

    ports = []

    for line in iface.splitlines():
        line = line.strip()

        if not line:
            continue

        if line.lower().startswith(("port", "name", "----")):
            continue

        name = line.split()[0]

        if name.startswith(exclude):
            continue

        if name.startswith(physical):
            ports.append(name)

    return len(ports)

# ============================================================
# ARUBA AOS-CX
# ============================================================
def get_aruba_cx_details(nc):
    details = {}

    details["hostname"] = nc.find_prompt().strip("# ").strip()

    snmp = send_command_safe(nc, "show snmp system", read_timeout=20)

    for line in snmp.splitlines():
        if "System location" in line:
            details["snmp_location"] = line.split(":", 1)[1].strip()

    prod = send_command_safe(nc, "show system | inc Product", read_timeout=20)

    if ":" in prod:
        details["model"] = prod.split(":", 1)[1].strip()

    if "model" not in details:
        system_output = send_command_safe(nc, "show system", read_timeout=20)

        for line in system_output.splitlines():
            if "Product" in line and ":" in line:
                details["model"] = line.split(":", 1)[1].strip()
                break

    iface = send_command_safe(nc, "show interfaces brief", read_timeout=30)

    ports = []

    for line in iface.splitlines():
        line = line.strip()

        if not line:
            continue

        parts = line.split()

        if not parts:
            continue

        first = parts[0]

        if "/" in first:
            ports.append(first)

    if ports:
        details["ports"] = len(set(ports))

    try:
        vsf = send_command_safe(nc, "show vsf", read_timeout=10)

        member_count = 0

        for line in vsf.splitlines():
            stripped = line.strip()

            if not stripped:
                continue

            # Aruba CX VSF output commonly has member rows beginning with numbers
            if stripped[0].isdigit():
                member_count += 1

        details["switch_num"] = member_count if member_count else 1

    except Exception:
        details["switch_num"] = 1

    return details

# ============================================================
# ARUBA OS / S2500
# ============================================================
def get_aruba_os_details(nc):
    details = {}

    try:
        if hasattr(nc, "check_enable_mode") and not nc.check_enable_mode():
            nc.enable()
    except Exception:
        logging.warning("Could not enter Aruba OS enable mode; continuing with available commands")

    prompt = nc.find_prompt().strip()
    details["hostname"] = prompt.strip("#> ").strip("()").strip()

    # Location - try likely Aruba OS commands
    location_outputs = [
        send_command_safe(nc, "show snmp-server system", read_timeout=20),
        send_command_safe(nc, "show snmp-server", read_timeout=20),
        send_command_safe(nc, "show snmp system", read_timeout=20),
    ]

    for output in location_outputs:
        for line in output.splitlines():
            lower = line.lower()

            if "location" in lower:
                if ":" in line:
                    details["snmp_location"] = line.split(":", 1)[1].strip()
                else:
                    details["snmp_location"] = line.strip()
                break

        if "snmp_location" in details:
            break

    # Model - try inventory/version/system outputs
    model_outputs = [
        send_command_safe(nc, "show inventory", read_timeout=30),
        send_command_safe(nc, "show version", read_timeout=30),
        send_command_safe(nc, "show system", read_timeout=30),
    ]

    for output in model_outputs:
        for line in output.splitlines():
            lower = line.lower()

            if "s2500" in lower or "s3500" in lower:
                details["model"] = line.strip()
                break

            if "model" in lower and ":" in line:
                details["model"] = line.split(":", 1)[1].strip()
                break

            if "product" in lower and ":" in line:
                details["model"] = line.split(":", 1)[1].strip()
                break

        if "model" in details:
            break

    # Fallback model from hostname, e.g. ae-s2500-48p-23-15-160
    if "model" not in details:
        hostname_lower = details["hostname"].lower()

        if "s2500-48p" in hostname_lower:
            details["model"] = "Aruba S2500-48P"
        elif "s2500-24p" in hostname_lower:
            details["model"] = "Aruba S2500-24P"
        elif "s2500" in hostname_lower:
            details["model"] = "Aruba S2500"
        elif "s3500" in hostname_lower:
            details["model"] = "Aruba S3500"

    # Ports - try to count from interface brief
    iface = (
        send_command_safe(nc, "show interface brief", read_timeout=30)
        or send_command_safe(nc, "show interfaces brief", read_timeout=30)
    )

    ports = []

    for line in iface.splitlines():
        line = line.strip()

        if not line:
            continue

        first = line.split()[0]
        first_lower = first.lower()

        if first_lower.startswith(
            (
                "ge",
                "gi",
                "fa",
                "te",
                "xg",
                "1/",
                "0/",
            )
        ):
            ports.append(first)

    if ports:
        details["ports"] = len(set(ports))
    else:
        # Fallback from hostname/model
        model_text = details.get("model", "") + " " + details.get("hostname", "")
        model_text = model_text.lower()

        if "48p" in model_text:
            details["ports"] = 48
        elif "24p" in model_text:
            details["ports"] = 24

    # S2500/S3500 treated as standalone unless you later add stacking logic
    details["switch_num"] = 1

    return details

# ============================================================
# CISCO IOS
# ============================================================
def get_cisco_ios_details(nc):
    details = {}

    try:
        nc.enable()
    except Exception:
        logging.warning("Could not enter Cisco enable mode; continuing with available commands")

    details["hostname"] = nc.find_prompt().strip("# ").strip()

    snmp = send_command_safe(nc, "show snmp location", read_timeout=20)

    if snmp.strip():
        details["snmp_location"] = snmp.strip()

    details["model"] = get_cisco_model_from_inventory(nc)
    details["ports"] = count_cisco_physical_ports(nc)

    sw = send_command_safe(nc, "show switch", read_timeout=30)

    members = [
        line for line in sw.splitlines()
        if line.strip().startswith(tuple(str(i) for i in range(1, 10)))
    ]

    details["switch_num"] = len(members) if members else 1

    return details

# ============================================================
# SWITCH HANDLER
# ============================================================
def get_switch_details(ip):
    platform_type = detect_platform(ip)
    session_log = os.path.join(LOG_DIR, f"session_{ip.replace('.', '_')}_{RUN_ID}.log")

    if platform_type == "aruba_cx":
        if not require_credential(SWITCH_USERNAME, "username", ip):
            return None

        if not require_credential(SWITCH_PASSWORD_AD, "password_AD or passwordAD", ip):
            return None

        device = {
            "device_type": "aruba_aoscx",
            "host": ip,
            "username": SWITCH_USERNAME,
            "password": SWITCH_PASSWORD_AD,
            "session_log": session_log,
            "conn_timeout": 30,
            "banner_timeout": 30,
            "auth_timeout": 30,
            "read_timeout_override": 30,
            "fast_cli": False,
        }

        platform_name = "Aruba CX"

    elif platform_type == "aruba_os":
        if not require_credential(SWITCH_USERNAME, "username", ip):
            return None

        if not require_credential(SWITCH_PASSWORD_ARUBA_OS, "aruba_pw", ip):
            return None

        device = {
            "device_type": "aruba_os",
            "host": ip,
            "username": SWITCH_USERNAME,
            "password": SWITCH_PASSWORD_ARUBA_OS,
            "secret": SWITCH_SECRET_ARUBA_OS,
            "session_log": session_log,
            "conn_timeout": 30,
            "banner_timeout": 30,
            "auth_timeout": 30,
            "read_timeout_override": 30,
            "global_delay_factor": 2,
            "fast_cli": False,
        }

        platform_name = "Aruba OS / S2500"

    else:
        if not require_credential(SWITCH_USERNAME, "username", ip):
            return None

        if not require_credential(SWITCH_PASSWORD, "Cisco password", ip):
            return None

        device = {
            "device_type": "cisco_ios",
            "host": ip,
            "username": SWITCH_USERNAME,
            "password": SWITCH_PASSWORD,
            "secret": SWITCH_SECRET,
            "session_log": session_log,
            "conn_timeout": 30,
            "banner_timeout": 30,
            "auth_timeout": 30,
            "read_timeout_override": 30,
            "fast_cli": False,
        }

        platform_name = "Cisco IOS"

    logging.info(f"[CONNECT] {ip} ({platform_name})")

    try:
        with ConnectHandler(**device) as nc:
            logging.info(f"[CONNECTED] {ip} ({platform_name})")

            if platform_type == "aruba_cx":
                return get_aruba_cx_details(nc)

            if platform_type == "aruba_os":
                return get_aruba_os_details(nc)

            return get_cisco_ios_details(nc)

    except Exception:
        logging.exception(f"[SSH FAIL] {ip} ({platform_name})")
        return None

# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Excel row number to start from",
    )

    parser.add_argument(
        "--ping-only",
        action="store_true",
        help="Ping only, no SSH",
    )

    parser.add_argument(
        "--ping-timeout",
        type=int,
        default=2,
        help="Ping timeout in seconds",
    )

    parser.add_argument(
        "--no-copy-back",
        action="store_true",
        help="Save locally only and do not copy back to OneDrive",
    )

    args = parser.parse_args()

    log_env_status()

    wb, ws = load_excel()
    cols = get_headers(ws)

    required_cols = [
        IP_COLUMN,
        HOSTNAME_COLUMN,
        LOCATION_COLUMN,
        SWITCH_NUM_COLUMN,
        MODEL_COLUMN,
        PORTS_COLUMN,
    ]

    missing = [col for col in required_cols if col not in cols]

    if missing:
        raise RuntimeError(f"Missing columns in Excel file: {missing}")

    ping_failures = []

    processed_count = 0
    ok_count = 0
    fail_count = 0
    ping_ok_count = 0
    ping_fail_count = 0

    for row_num in range(args.start_row, ws.max_row + 1):
        ip = ws.cell(row=row_num, column=cols[IP_COLUMN]).value

        if not ip:
            continue

        ip = str(ip).strip()

        if not ip:
            continue

        processed_count += 1

        logging.info(f"[ROW {row_num}] Starting {ip}")

        if args.ping_only:
            ping_success = ping_host(ip, args.ping_timeout)

            if ping_success:
                ping_ok_count += 1

                log_row_result(
                    row_num=row_num,
                    ip=ip,
                    status="PING OK",
                )
            else:
                ping_fail_count += 1
                ping_failures.append(ip)
                highlight_cell(ws, row_num, cols[IP_COLUMN])

                log_row_result(
                    row_num=row_num,
                    ip=ip,
                    status="PING FAIL",
                    reason="Host did not respond to ping",
                )

            continue

        data = get_switch_details(ip)

        if not data:
            fail_count += 1
            highlight_cell(ws, row_num, cols[IP_COLUMN])

            log_row_result(
                row_num=row_num,
                ip=ip,
                status="SSH FAIL",
                reason="No data returned from switch",
            )

            continue

        field_map = {
            "hostname": HOSTNAME_COLUMN,
            "snmp_location": LOCATION_COLUMN,
            "switch_num": SWITCH_NUM_COLUMN,
            "model": MODEL_COLUMN,
            "ports": PORTS_COLUMN,
        }

        for field, column_name in field_map.items():
            if field in data:
                update_cell(ws, row_num, cols[column_name], data[field])

        ok_count += 1

        log_row_result(
            row_num=row_num,
            ip=ip,
            status="OK",
            data=data,
        )

    if args.ping_only:
        logging.info(
            f"[SUMMARY] rows_processed={processed_count} "
            f"ping_ok={ping_ok_count} "
            f"ping_failed={ping_fail_count}"
        )
    else:
        logging.info(
            f"[SUMMARY] rows_processed={processed_count} "
            f"ssh_ok={ok_count} "
            f"ssh_failed={fail_count}"
        )

    if args.ping_only and ping_failures:
        with open(PING_FAIL_LOG, "w") as f:
            for ip in ping_failures:
                f.write(f"{ip}\n")

        logging.warning(f"Ping failures written to {PING_FAIL_LOG}")

    # Save local copy once at the end
    wb.save(LOCAL_EXCEL_FILE)
    wb.close()

    logging.info(f"Local workbook saved: {LOCAL_EXCEL_FILE}")

    if args.no_copy_back:
        logging.warning(
            f"--no-copy-back used. Workbook was saved locally only: {LOCAL_EXCEL_FILE}"
        )
    else:
        copied_back = copy_result_back_to_onedrive()

        if not copied_back:
            logging.warning(
                "Copy-back failed or was skipped. "
                f"Use this edited local file manually: {LOCAL_EXCEL_FILE}"
            )

    logging.info("=== Script complete ===")


if __name__ == "__main__":
    main()