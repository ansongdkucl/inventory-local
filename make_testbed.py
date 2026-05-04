#!/usr/bin/env python3

import re
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook


# ============================================================
# Settings
# ============================================================

SCRIPT_DIR = Path(__file__).resolve().parent

INPUT_FILE = SCRIPT_DIR / "inventory.xlsx"
OUTPUT_FILE = SCRIPT_DIR / "inventory/hosts.yaml"

DEFAULT_PORT = 22
DEFAULT_GROUP = "switches"


# ============================================================
# Column name matching
# ============================================================

COLUMN_ALIASES = {
    "hostname": [
        "hostname",
        "host",
        "device",
        "device name",
        "switch",
        "switch name",
        "name",
    ],
    "ip_address": [
        "ip address",
        "ip",
        "management ip",
        "mgmt ip",
        "management address",
        "host ip",
    ],
    "vendor": [
        "vendor",
        "platform",
        "device type",
        "os",
    ],
    "groups": [
        "groups",
        "group",
        "nornir group",
    ],
    "building": [
        "building",
        "building number",
        "site",
    ],
    "cab": [
        "cab",
        "cabinet",
        "rack",
    ],
}


# ============================================================
# Vendor to Nornir platform mapping
# ============================================================

VENDOR_PLATFORM_MAP = {
    "cisco": "cisco_ios",
    "cisco ios": "cisco_ios",
    "ios": "cisco_ios",
    "cisco-ios": "cisco_ios",
    "cisco_os": "cisco_ios",
    "cisco_os_ios": "cisco_ios",

    "aruba": "aruba_osswitch",
    "aruba os": "aruba_osswitch",
    "aruba-os": "aruba_osswitch",
    "aruba_os": "aruba_osswitch",
    "aruba procurve": "aruba_osswitch",
    "procurve": "aruba_osswitch",

    "hp": "hp_procurve",
    "hpe": "hp_procurve",

    "aruba cx": "aruba_aoscx",
    "aoscx": "aruba_aoscx",
    "aruba_aoscx": "aruba_aoscx",
}

def normalise_header(value):
    """
    Normalise spreadsheet column headings so matching is easier.
    Example:
        'IP Address ' -> 'ip address'
    """
    if value is None:
        return ""

    value = str(value).strip().lower()
    value = re.sub(r"\s+", " ", value)
    return value


def find_columns(header_row):
    """
    Finds the spreadsheet column numbers for required fields.
    Returns a dictionary like:
        {
            "hostname": 1,
            "ip_address": 2,
            ...
        }
    """
    normalised_headers = {
        normalise_header(cell.value): index
        for index, cell in enumerate(header_row, start=1)
        if cell.value is not None
    }

    found = {}

    for target_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias = normalise_header(alias)

            if alias in normalised_headers:
                found[target_name] = normalised_headers[alias]
                break

    return found


def get_cell_value(sheet, row_number, column_number):
    if not column_number:
        return None

    value = sheet.cell(row=row_number, column=column_number).value

    if value is None:
        return None

    value = str(value).strip()

    if value == "":
        return None

    return value


def make_host_key(hostname, ip_address):
    """
    Nornir host key.

    Priority:
      1. Use the hostname column if present.
      2. Otherwise create one from the IP address.

    Example:
        172.18.0.70 -> ce18-0-70
    """
    if hostname:
        return hostname.strip()

    if ip_address:
        parts = ip_address.strip().split(".")

        if len(parts) == 4:
            return f"ce{parts[1]}-{parts[2]}-{parts[3]}"

    return None


def map_vendor_to_platform(vendor):
    """
    Converts spreadsheet vendor values to Nornir platform values.
    """
    if not vendor:
        return "cisco-ios"

    vendor_clean = vendor.strip().lower()

    return VENDOR_PLATFORM_MAP.get(vendor_clean, vendor_clean)


def parse_groups(groups_value):
    """
    Supports:
        switches
        switches,cisco
        switches;building-444
    """
    if not groups_value:
        return [DEFAULT_GROUP]

    groups = re.split(r"[;,]", groups_value)

    cleaned_groups = [
        group.strip()
        for group in groups
        if group.strip()
    ]

    if not cleaned_groups:
        return [DEFAULT_GROUP]

    return cleaned_groups


def main():
    if not INPUT_FILE.exists():
        print(f"ERROR: Could not find {INPUT_FILE}")
        sys.exit(1)

    workbook = load_workbook(INPUT_FILE, data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    columns = find_columns(header_row)

    required_columns = ["ip_address"]

    missing = [
        column
        for column in required_columns
        if column not in columns
    ]

    if missing:
        print("ERROR: Missing required spreadsheet columns:")
        for column in missing:
            print(f"  - {column}")

        print("\nDetected columns:")
        for cell in header_row:
            if cell.value:
                print(f"  - {cell.value}")

        sys.exit(1)

    hosts = {}

    for row_number in range(2, sheet.max_row + 1):
        hostname_value = get_cell_value(
            sheet,
            row_number,
            columns.get("hostname"),
        )

        ip_address = get_cell_value(
            sheet,
            row_number,
            columns.get("ip_address"),
        )

        vendor = get_cell_value(
            sheet,
            row_number,
            columns.get("vendor"),
        )

        groups_value = get_cell_value(
            sheet,
            row_number,
            columns.get("groups"),
        )

        building = get_cell_value(
            sheet,
            row_number,
            columns.get("building"),
        )

        cab = get_cell_value(
            sheet,
            row_number,
            columns.get("cab"),
        )

        if not ip_address:
            print(f"Skipping row {row_number}: no IP address")
            continue

        host_key = make_host_key(hostname_value, ip_address)

        if not host_key:
            print(f"Skipping row {row_number}: could not create host key")
            continue

        platform = map_vendor_to_platform(vendor)
        groups = parse_groups(groups_value)

        host_data = {
            "hostname": ip_address,
            "port": DEFAULT_PORT,
            "platform": platform,
            "groups": groups,
            "data": {},
        }

        if building:
            host_data["data"]["building"] = building

        if cab:
            host_data["data"]["cab"] = cab

        hosts[host_key] = host_data

    with open(OUTPUT_FILE, "w", encoding="utf-8") as file:
        yaml.safe_dump(
            hosts,
            file,
            sort_keys=False,
            default_flow_style=False,
            allow_unicode=True,
        )

    print(f"Created {OUTPUT_FILE}")
    print(f"Total hosts written: {len(hosts)}")


if __name__ == "__main__":
    main()