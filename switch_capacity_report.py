#!/usr/bin/env python3

import argparse
import os
import re
from datetime import datetime
from pathlib import Path

from nornir import InitNornir
from nornir.core.inventory import ConnectionOptions
from nornir_netmiko.tasks import netmiko_send_command
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass


PHYS_PORT_RE = re.compile(
    r"^(gi|gig|gigabitethernet|fa|fastethernet|te|ten|tengigabitethernet|"
    r"tw|twe|fo|hu|eth|ethernet|\d+/\d+/\d+)",
    re.I,
)

OUTLET_RE = re.compile(r"\b\d{1,4}/\d{1,4}/\d{1,4}\b")


def is_cisco(platform):
    p = (platform or "").lower()
    return "cisco" in p or "ios" in p


def is_old_aruba(host):
    p = (host.platform or "").lower()
    ip = str(host.hostname or "")

    return (
        "aruba_os" in p
        or "aruba-osswitch" in p
        or "aruba_osswitch" in p
        or ip.startswith("172.23.")
        or ip.startswith("172.30.30.")
    )


def netmiko_platform(platform):
    p = (platform or "").lower().strip()

    platform_map = {
        "cisco-ios": "cisco_ios",
        "cisco_ios": "cisco_ios",
        "ios": "cisco_ios",
        "aruba_aoscx": "aruba_aoscx",
        "aruba-aoscx": "aruba_aoscx",
        "aruba_cx": "aruba_aoscx",
        "aoscx": "aruba_aoscx",
        "aruba_os": "aruba_os",
        "aruba-osswitch": "aruba_os",
        "aruba_osswitch": "aruba_os",
    }

    return platform_map.get(p, platform)


def setup_connection(task):
    host = task.host
    username = os.getenv("username")

    if is_cisco(host.platform):
        password = os.getenv("password")
        secret = os.getenv("secret") or password
    elif is_old_aruba(host):
        password = os.getenv("aruba_pw")
        secret = None
    else:
        password = os.getenv("passwordAD")
        secret = None

    if not username:
        raise RuntimeError("Missing environment variable: username")

    if not password:
        raise RuntimeError(
            f"{host.name}: missing password. "
            "Cisco uses password, Aruba CX uses passwordAD, old Aruba uses aruba_pw."
        )

    extras = {
        "conn_timeout": 30,
        "auth_timeout": 30,
        "banner_timeout": 30,
        "session_timeout": 60,
        "fast_cli": False,
    }

    if secret:
        extras["secret"] = secret

    host.connection_options["netmiko"] = ConnectionOptions(
        hostname=host.hostname,
        port=host.port or 22,
        username=username,
        password=password,
        platform=netmiko_platform(host.platform),
        extras=extras,
    )


def is_physical_port(port):
    p = port.lower()
    return bool(PHYS_PORT_RE.match(p)) and not p.startswith(
        ("po", "vlan", "lo", "mgmt", "null")
    )


def cab_label(host):
    building = str(host.data.get("building", "")).strip()
    cab = str(host.data.get("cab", host.data.get("cabinet", ""))).strip()

    if "/" in cab:
        return cab

    return f"{building}/{cab}".strip("/")


def parse_user_age(value):
    if not value:
        return None

    value = value.strip().lower()
    value = value.replace("months", "mo").replace("month", "mo")

    m = re.fullmatch(
        r"(\d+)\s*(y|yr|year|years|mo|m|w|week|weeks|d|day|days|h|hour|hours)",
        value,
    )

    if not m:
        raise argparse.ArgumentTypeError(
            f"Invalid age '{value}'. Use examples like 1w, 4w, 3mo, 6m, 90d."
        )

    number = int(m.group(1))
    unit = m.group(2)

    sizes = {
        "y": 365 * 86400,
        "yr": 365 * 86400,
        "year": 365 * 86400,
        "years": 365 * 86400,
        "mo": 30 * 86400,
        "m": 30 * 86400,
        "w": 7 * 86400,
        "week": 7 * 86400,
        "weeks": 7 * 86400,
        "d": 86400,
        "day": 86400,
        "days": 86400,
        "h": 3600,
        "hour": 3600,
        "hours": 3600,
    }

    return number * sizes[unit]


def parse_age(text):
    if not text:
        return None

    t = text.lower()
    t = t.replace(",", " ")
    t = t.replace("ago", " ")
    t = t.replace("(s)", "s")
    t = t.strip()

    if any(x in t for x in ["never", "unknown", "n/a", "--"]):
        return None

    t = re.sub(r".*uptime\s+is\s+", "", t)
    t = re.sub(r".*up\s*time\s*[:=]\s*", "", t)
    t = re.sub(r".*last\s+link\s+flapped\s+", "", t)
    t = re.sub(r".*last\s+link\s+state\s+change\s*[:=]?\s*", "", t)
    t = re.sub(r".*last\s+input\s+", "", t)
    t = re.sub(r"(?<=[a-z])(?=\d)", " ", t)

    hm = re.search(r"\b(\d{1,2}):(\d{2}):(\d{2})\b", t)
    if hm:
        return int(hm.group(1)) * 3600 + int(hm.group(2)) * 60 + int(hm.group(3))

    sizes = {
        "year": 365 * 86400,
        "years": 365 * 86400,
        "yr": 365 * 86400,
        "yrs": 365 * 86400,
        "y": 365 * 86400,
        "month": 30 * 86400,
        "months": 30 * 86400,
        "mo": 30 * 86400,
        "week": 7 * 86400,
        "weeks": 7 * 86400,
        "w": 7 * 86400,
        "day": 86400,
        "days": 86400,
        "d": 86400,
        "hour": 3600,
        "hours": 3600,
        "hr": 3600,
        "hrs": 3600,
        "h": 3600,
        "minute": 60,
        "minutes": 60,
        "min": 60,
        "mins": 60,
        "second": 1,
        "seconds": 1,
        "sec": 1,
        "secs": 1,
        "s": 1,
    }

    total = 0

    for number, unit in re.findall(
        r"(\d+)\s*(years?|yrs?|y|months?|mo|weeks?|w|days?|d|hours?|hrs?|h|minutes?|mins?|min|seconds?|secs?|s)\b",
        t,
    ):
        total += int(number) * sizes[unit]

    return total or None


def fmt_age(seconds, ago=False):
    if seconds is None:
        return "unknown"

    for name, size in [
        ("year", 365 * 86400),
        ("month", 30 * 86400),
        ("week", 7 * 86400),
        ("day", 86400),
        ("hour", 3600),
        ("minute", 60),
    ]:
        if seconds >= size:
            value = int(seconds // size)
            text = f"{value} {name}{'' if value == 1 else 's'}"
            return f"{text} ago" if ago else text

    return "less than 1 minute ago" if ago else "less than 1 minute"


def run_show(task, command):
    result = task.run(
        task=netmiko_send_command,
        command_string=command,
        enable=is_cisco(task.host.platform),
        read_timeout=60,
    )
    return str(result.result or "")


def uptime_cmd(platform):
    return "show version | include uptime" if is_cisco(platform) else "show system"


def status_cmd(platform):
    return "show interfaces status" if is_cisco(platform) else "show interface brief"


def detail_cmd(platform, port):
    return f"show interfaces {port}" if is_cisco(platform) else f"show interface {port}"


def config_int_cmd(port):
    return f"show running-config interface {port}"


def parse_uptime(output):
    for line in output.splitlines():
        if re.search(r"\buptime\b|up\s*time", line, re.I):
            age = parse_age(line)
            if age:
                return age

    return parse_age(output)


def parse_snmp_location(output):
    for line in output.splitlines():
        line = line.strip()

        m = re.search(r"snmp-server\s+(?:system-)?location\s+(.+)", line, re.I)
        if m:
            return m.group(1).strip()

    return ""


def get_snmp_location(task):
    output = run_show(task, "show running-config | include snmp")
    location = parse_snmp_location(output)

    return location or str(task.host.data.get("snmp_location", "")).strip()


def vlan_from_cisco_status(line):
    parts = line.split()

    for status in ["connected", "notconnect", "disabled", "err-disabled", "inactive"]:
        if status in parts:
            idx = parts.index(status)
            if len(parts) > idx + 1:
                return parts[idx + 1]

    return ""


def candidate_ports(platform, output, include_admin_down=False):
    ports = []

    for raw_line in output.splitlines():
        line = raw_line.strip()

        if not line or line.lower().startswith(("port", "interface", "----")):
            continue

        port = line.split()[0]

        if not is_physical_port(port):
            continue

        low = line.lower()

        if is_cisco(platform):
            if "notconnect" in low or (include_admin_down and "disabled" in low):
                ports.append(
                    {
                        "port": port,
                        "vlan": vlan_from_cisco_status(line),
                    }
                )
        else:
            if re.search(r"\bdown\b", low):
                ports.append(
                    {
                        "port": port,
                        "vlan": "",
                    }
                )

    return ports


def parse_vlan_from_config(output):
    if re.search(r"^\s*switchport\s+mode\s+trunk\b", output, re.I | re.M):
        return "trunk"

    if re.search(r"^\s*vlan\s+trunk\s+allowed\b", output, re.I | re.M):
        return "trunk"

    patterns = [
        r"^\s*switchport\s+access\s+vlan\s+(\d+)",
        r"^\s*vlan\s+access\s+(\d+)",
        r"^\s*untagged\s+vlan\s+(\d+)",
    ]

    for pattern in patterns:
        m = re.search(pattern, output, re.I | re.M)
        if m:
            return m.group(1)

    return ""


def parse_description(output):
    patterns = [
        r"^\s*Description\s*[:=]\s*(.+)$",
        r"^\s*description\s+(.+)$",
    ]

    for pattern in patterns:
        m = re.search(pattern, output, re.I | re.M)
        if m:
            return m.group(1).strip()

    return ""


def outlet_from_description(description):
    m = OUTLET_RE.search(description or "")
    return m.group(0) if m else ""


def last_used_from_detail(output):
    patterns = [
        r"Last\s+link\s+flapped\s+([^\r\n]+)",
        r"Last\s+link\s+state\s+change\s*[:=]?\s*([^\r\n]+)",
        r"Last\s+input\s+([^,\r\n]+)",
    ]

    for pattern in patterns:
        m = re.search(pattern, output, re.I)
        if m:
            return parse_age(m.group(1))

    return None


def is_admin_down(output):
    output = output.lower()

    return (
        "administratively down" in output
        or re.search(r"admin\s+state\s*(is|:)?\s*down", output) is not None
    )


def is_connected(output):
    output = output.lower()

    return (
        "line protocol is up" in output
        or re.search(r"link\s+state\s*(is|:)?\s*up", output) is not None
    )


def is_trunk_or_routed(vlan):
    vlan = (vlan or "").lower()
    return vlan in ["trunk", "routed"] or "trunk" in vlan


def collect_switch_capacity(task, requested_age, include_admin_down, include_trunks):
    setup_connection(task)

    host = task.host
    platform = host.platform or ""
    hostname = host.name
    cab = cab_label(host)

    rows = []
    warnings = []

    try:
        uptime = parse_uptime(run_show(task, uptime_cmd(platform)))
    except Exception as e:
        uptime = None
        warnings.append(f"{hostname}: failed to read uptime: {e}")

    try:
        snmp_location = get_snmp_location(task)
    except Exception:
        snmp_location = str(host.data.get("snmp_location", "")).strip()

    effective_age = min(requested_age, uptime) if requested_age and uptime else requested_age

    try:
        status_output = run_show(task, status_cmd(platform))
        ports = candidate_ports(platform, status_output, include_admin_down)
    except Exception as e:
        return {
            "rows": rows,
            "summary": {
                "hostname": hostname,
                "uptime": uptime,
                "effective_age": effective_age,
            },
            "warnings": [f"{hostname}: failed to read interface status: {e}"],
        }

    for item in ports:
        port = item["port"]

        try:
            detail_output = run_show(task, detail_cmd(platform, port))
            config_output = run_show(task, config_int_cmd(port))
        except Exception as e:
            warnings.append(f"{hostname} {port}: failed to read port details: {e}")
            continue

        if is_connected(detail_output):
            continue

        if is_admin_down(detail_output) and not include_admin_down:
            continue

        vlan = item.get("vlan") or parse_vlan_from_config(config_output)

        if is_trunk_or_routed(vlan) and not include_trunks:
            continue

        port_age = last_used_from_detail(detail_output)

        if port_age is None:
            port_age = uptime

        if port_age and uptime:
            port_age = min(port_age, uptime)

        if requested_age:
            if effective_age is None or port_age is None:
                continue

            if port_age < effective_age:
                continue

        description = parse_description(detail_output) or parse_description(config_output)

        rows.append(
            {
                "cab": cab,
                "hostname": hostname,
                "port": port,
                "vlan": vlan,
                "outlet": outlet_from_description(description),
                "snmp_location": snmp_location,
                "last_used": fmt_age(port_age, ago=True),
                "uptime": fmt_age(uptime),
            }
        )

    return {
        "rows": rows,
        "summary": {
            "hostname": hostname,
            "uptime": uptime,
            "effective_age": effective_age,
        },
        "warnings": warnings,
    }


def autosize_columns(sheet, max_width):
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        width = 0

        for row_idx in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            width = max(width, len(str(value or "")))

        sheet.column_dimensions[col_letter].width = min(width + 2, max_width)


def write_xlsx(rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Switch Capacity"

    headers = [
        "CAB",
        "Hostname",
        "port",
        "Vlan",
        "outlet",
        "snmp location",
        "last used",
        "uptime",
    ]

    ws.append(["Switch Capacity Report"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(headers)

    for row in rows:
        ws.append(
            [
                row["cab"],
                row["hostname"],
                row["port"],
                row["vlan"],
                row["outlet"],
                row["snmp_location"],
                row["last_used"],
                row["uptime"],
            ]
        )

    ws.append([])
    ws.append([])
    ws.append(["Total Free ports", "", len(rows), "", "", "", "", ""])

    thin = Side(style="thin")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top")

    for cell in ws[3]:
        cell.font = Font(bold=True)

    total_row = ws.max_row
    ws.cell(total_row, 1).font = Font(bold=True)
    ws.cell(total_row, 3).font = Font(bold=True)

    autosize_columns(ws, 55)
    wb.save(filename)


def main():
    parser = argparse.ArgumentParser(
        description="Create a switch capacity report for free ports in a CAB"
    )

    parser.add_argument(
        "--config",
        default="config.yaml",
        help="Nornir config file. Default: config.yaml",
    )

    parser.add_argument(
        "--target",
        required=True,
        help="CAB target from Nornir data, for example 86/1 or 86/100",
    )

    parser.add_argument(
        "--unused-for",
        type=parse_user_age,
        help="Requested unused period, for example 1w, 4w, 3mo, 6m",
    )

    parser.add_argument(
        "--include-admin-down",
        action="store_true",
        help="Include administratively disabled ports",
    )

    parser.add_argument(
        "--include-trunks",
        action="store_true",
        help="Include trunk or routed ports",
    )

    parser.add_argument(
        "--xlsx",
        help="Output Excel filename",
    )

    args = parser.parse_args()

    nr = InitNornir(config_file=args.config)
    target = nr.filter(filter_func=lambda h: cab_label(h) == args.target)

    if not target.inventory.hosts:
        raise SystemExit(f"No switches found in Nornir inventory for CAB {args.target}")

    result = target.run(
        task=collect_switch_capacity,
        requested_age=args.unused_for,
        include_admin_down=args.include_admin_down,
        include_trunks=args.include_trunks,
    )

    rows = []
    summaries = []
    warnings = []

    for host_result in result.values():
        if host_result.failed:
            warnings.append(str(host_result.exception))
            continue

        data = host_result.result or {}
        rows.extend(data.get("rows", []))
        summaries.append(data.get("summary", {}))
        warnings.extend(data.get("warnings", []))

    rows.sort(key=lambda r: (r["cab"], r["hostname"], r["port"]))

    outfile = args.xlsx or f"capacity_{args.target.replace('/', '_')}_{datetime.now():%Y%m%d_%H%M}.xlsx"

    write_xlsx(rows, outfile)

    print("\nSwitch Capacity Report")
    print("-" * 70)
    print(f"CAB: {args.target}")
    print(f"Requested unused period: {fmt_age(args.unused_for) if args.unused_for else 'all free ports'}")
    print(f"Total free ports found: {len(rows)}")
    print(f"Report written: {Path(outfile).resolve()}")

    print("\nSwitch uptime / effective report window")
    print("-" * 70)

    for summary in summaries:
        print(
            f"{summary['hostname']:<25} "
            f"uptime={fmt_age(summary['uptime']):<15} "
            f"effective-window={fmt_age(summary['effective_age']) if summary['effective_age'] else 'not set'}"
        )

    if warnings:
        print("\nWarnings")
        print("-" * 70)
        for warning in warnings:
            print(f"- {warning}")


if __name__ == "__main__":
    main()